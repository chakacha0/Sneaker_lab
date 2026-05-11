from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from app.repositories.orders_repo import (
    create_order,
    get_order_by_id,
    get_user_orders,
    calculate_order_total_with_promo,
    get_sales_statistics,
    get_all_orders_admin,
    update_order_status,
)
from app.repositories.cart_repo import get_cart_items, get_or_create_cart
from datetime import datetime
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/orders")

@router.post("/")
def create_order_endpoint(data: dict):
    """
    Создает заказ на основе корзины пользователя
    Требует: user_id, address_id
    Опционально: promo_code
    """
    user_id = data.get("user_id")
    address_id = data.get("address_id")
    promo_code = data.get("promo_code")
    
    if not user_id or not address_id:
        raise HTTPException(status_code=400, detail="user_id и address_id обязательны")
    
    try:
        order = create_order(
            user_id=user_id,
            address_id=address_id,
            promo_code=promo_code
        )
        return {
            "message": "Заказ успешно создан",
            "order": order
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка создания заказа: {str(e)}")


@router.get("/admin/list")
def list_all_orders_admin():
    """All orders (admin panel)."""
    try:
        return get_all_orders_admin()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/admin/{order_id}/status")
def patch_order_status(order_id: int, data: dict):
    """Update order status."""
    status = data.get("status")
    if not status or not isinstance(status, str):
        raise HTTPException(status_code=400, detail="Field status is required")
    try:
        updated = update_order_status(order_id, status.strip())
        if not updated:
            raise HTTPException(status_code=404, detail="Order not found")
        return {"message": "Status updated", "order": updated}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{order_id}")
def get_order(order_id: int):
    """
    Получает заказ по ID
    """
    try:
        order = get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Заказ не найден")
        return order
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/user/{user_id}")
def get_orders_by_user(user_id: int):
    """
    Получает все заказы пользователя
    """
    try:
        orders = get_user_orders(user_id)
        return orders
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/calculate-total")
def calculate_order_total(data: dict):
    """
    Вычисляет итоговую стоимость заказа с учетом промокода
    Требует: user_id, promo_code (опционально)
    """
    user_id = data.get("user_id")
    promo_code = data.get("promo_code")
    
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id обязателен")
    
    try:
        # Получаем корзину пользователя
        cart = get_or_create_cart(user_id)
        cart_id = cart.get("cart_id")
        
        # Получаем товары из корзины
        cart_items = get_cart_items(cart_id)
        
        if not cart_items:
            raise HTTPException(status_code=400, detail="Корзина пуста")
        
        # Вычисляем исходную стоимость заказа
        original_total = sum(item.get("price", 0) * item.get("quantity", 0) for item in cart_items)
        
        # Вычисляем итоговую стоимость с учетом промокода
        final_total, promo_valid, promo_message = calculate_order_total_with_promo(
            original_total,
            promo_code.strip().upper() if promo_code and promo_code.strip() else None
        )
        
        # Убеждаемся, что все значения приведены к float
        original_total_float = float(original_total)
        final_total_float = float(final_total)
        
        return {
            "original_total": original_total_float,
            "final_total": final_total_float,
            "promo_code": promo_code.strip().upper() if promo_code and promo_code.strip() else None,
            "promo_valid": promo_valid,
            "promo_message": promo_message,
            "discount": max(0, original_total_float - final_total_float)  # Скидка не может быть отрицательной
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/reports/sales")
def get_sales_report(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)")
):
    """
    Получает статистику продаж за указанный период
    """
    try:
        stats = get_sales_statistics(start_date, end_date)
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка получения статистики: {str(e)}")

@router.get("/reports/sales/excel")
def generate_sales_excel_report(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)")
):
    """
    Генерирует Excel отчет о продажах
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500, 
            detail="Библиотека openpyxl не установлена. Установите её: pip install openpyxl"
        )
    
    try:
        # Получаем статистику
        stats = get_sales_statistics(start_date, end_date)
        
        # Проверяем, что статистика получена
        if not stats:
            raise HTTPException(status_code=500, detail="Не удалось получить статистику продаж")
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Стили
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовок отчета
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Sales Report"
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        
        # Период отчета
        period_text = "All time"
        if start_date and end_date:
            period_text = f"From {start_date} to {end_date}"
        elif start_date:
            period_text = f"From {start_date}"
        elif end_date:
            period_text = f"Until {end_date}"
        
        ws.merge_cells('A2:D2')
        period_cell = ws['A2']
        period_cell.value = period_text
        period_cell.alignment = center_alignment
        
        # Дата генерации
        ws.merge_cells('A3:D3')
        date_cell = ws['A3']
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = center_alignment
        
        # Пустая строка
        ws.append([])
        
        # Общая статистика
        row = 5
        ws[f'A{row}'] = "Total Orders:"
        ws[f'B{row}'] = stats['total_orders']
        ws[f'A{row}'].font = bold_font
        
        row += 1
        ws[f'A{row}'] = "Total Items Sold:"
        ws[f'B{row}'] = stats['total_items_sold']
        ws[f'A{row}'].font = bold_font
        
        row += 1
        ws[f'A{row}'] = "Total Revenue:"
        ws[f'B{row}'] = f"{stats['total_revenue']:.2f} $"
        ws[f'A{row}'].font = bold_font
        
        # Пустая строка
        row += 2
        ws.append([])
        
        # Заголовки таблицы заказов
        row += 1
        headers = ["Order ID", "Date", "Customer", "Items", "Total Price"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Данные заказов
        for order in stats['orders']:
            row += 1
            order_id = order.get('order_id')
            order_date = order.get('created_at')
            if isinstance(order_date, str):
                try:
                    order_date = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
                except:
                    pass
            elif not hasattr(order_date, 'strftime'):
                order_date = datetime.now()
            
            customer_name = ""
            if order.get('first_name') or order.get('last_name'):
                customer_name = f"{order.get('first_name', '')} {order.get('last_name', '')}".strip()
            if not customer_name and order.get('email'):
                customer_name = order.get('email')
            
            items_count = sum(item.get('quantity', 0) for item in order.get('items', []))
            total_price = float(order.get('total_price', 0))
            
            ws.cell(row=row, column=1).value = order_id
            ws.cell(row=row, column=2).value = order_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(order_date, 'strftime') else str(order_date)
            ws.cell(row=row, column=3).value = customer_name
            ws.cell(row=row, column=4).value = items_count
            ws.cell(row=row, column=5).value = f"{total_price:.2f} $"
        
        # Автоматическая ширина колонок
        # Используем координаты напрямую, чтобы избежать проблем с объединенными ячейками
        from openpyxl.utils import get_column_letter
        
        # Определяем максимальную колонку (E = 5 для наших 5 колонок)
        max_col = 5
        for col_num in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            # Проходим по всем строкам в этой колонке
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    # Пропускаем объединенные ячейки (они не имеют value напрямую)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                except (AttributeError, TypeError):
                    pass
            
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Формируем имя файла
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = str(e)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчета: {error_detail}")
